from __future__ import annotations

import io

import pytest
from openpyxl import load_workbook

from tpt_api.datahub import batch_add_tags, delete_tags_physical, export_tags, list_tags
from tpt_api.types import DataTypes, TagTypes

from tests.support.cleanup import delete_tag_if_exists
from tests.support.polling import wait_until
from tests.support.rt_helpers import get_rt_point
from tests.support.ua2_helpers import (
    browse_entry_to_batch_info,
    pick_unused_nodes,
    setup_ds_and_tag,
    setup_ds_only,
    teardown_ds_tag_mocker,
    wait_ds_alive,
)
from tests.support.ua2_import_helpers import (
    assert_tag_fields,
    build_import_workbook,
    export_template_bytes,
    import_bytes,
)


def _records_for_names(api, names: list[str]) -> list[dict]:
    wanted = set(names)
    page = list_tags(api, page=1, page_size=max(100, len(names) + 20))
    return [r for r in page.get("records") or [] if r.get("tagName") in wanted]


def _created_ids(api, names: list[str]) -> list[int]:
    return [int(r["id"]) for r in _records_for_names(api, names)]


def _batch_infos(nodes: list[dict], ds_id: int, prefix: str, *, unit: str = "", desc: str = "") -> list[dict]:
    return [
        browse_entry_to_batch_info(
            node,
            ds_id=ds_id,
            tag_name=f"{prefix}_{index:03d}",
            only_read=True,
            unit=unit,
            tag_desc=desc,
        )
        for index, node in enumerate(nodes)
    ]


def _wait_rt(api, tag_name: str) -> None:
    wait_until(
        f"rt:{tag_name}",
        lambda: (
            (point := get_rt_point(api, tag_name)).get("tagValue") is not None
            and point.get("quality") not in (None, 0)
        ),
        timeout=60.0,
        interval=0.5,
    )


@pytest.mark.case(
    id="UA-2-3-028", chapter="UA-2-3", title="往返_导出修改导入再导出",
    preconditions=["目标 datasource alive"],
    steps=["导出", "修改 Unit、Frequency、Description", "覆盖导入", "再次导出并比较"],
    expected=["修改字段生效", "未修改字段保留", "RT 仍可用"],
)
@pytest.mark.integration
@pytest.mark.destructive
def test_tag_import_roundtrip_export_modify_import_export(api, settings, tmp_path_factory, mocker_endpoint):
    ctx = setup_ds_and_tag(
        api, settings, mocker_endpoint, tmp_path_factory, "UA-2-3-028",
        tag_base_name="2_ua23_028_roundtrip_1", data_type=DataTypes["DOUBLE"],
        tag_type=TagTypes["一次位号"], only_read=False, unit="kPa",
        tag_desc="original description", frequency=10,
        nodes=[{"name": "ua23_028_roundtrip_", "type": "Double", "default": 12.5, "writable": True}],
        namespace_index=2, cycle=500,
    )
    try:
        original = assert_tag_fields(api, ctx["tag_name"])
        first_export = export_template_bytes(api, ctx["tag_id"])
        wb = load_workbook(io.BytesIO(first_export))
        header = [cell.value for cell in wb.active[1]]
        wb.close()
        row = {
            "Tag Name": ctx["tag_name"],
            "Base Tag Name": original["tagBaseName"],
            "Tag Type": "Primary Tag",
            "Datasource Name": ctx["ds_name"],
            "Unit": "MHz",
            "Data Type": "DOUBLE",
            "Frequency": 25,
            "Description": "roundtrip modified",
            "Group Name": original.get("groupName") or "Root",
            "Real-time Push": "true",
            "Readonly": "false",
        }
        assert "Unit" in header and "Frequency" in header and "Description" in header
        import_bytes(api, build_import_workbook(first_export, [row]), conflict_strategy=1, tmp_path_factory=tmp_path_factory)
        updated = assert_tag_fields(
            api, ctx["tag_name"], dsId=ctx["ds_id"], tagBaseName=original["tagBaseName"],
            dataType=original["dataType"], unit="MHz", frequency=25, tagDesc="roundtrip modified",
        )
        assert updated["id"] == original["id"]
        second_export = export_template_bytes(api, int(updated["id"]))
        first_wb = load_workbook(io.BytesIO(first_export), data_only=True)
        second_wb = load_workbook(io.BytesIO(second_export), data_only=True)
        first_row = list(first_wb.active.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        second_row = list(second_wb.active.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        first_wb.close()
        second_wb.close()
        for index in (0, 1, 2, 3, 5, 16, 17, 18):
            assert second_row[index] == first_row[index], f"unchanged export column {index} changed"
        assert second_row[4] == "MHz" and second_row[8] == 25 and second_row[15] == "roundtrip modified"
        _wait_rt(api, ctx["tag_name"])
    finally:
        teardown_ds_tag_mocker(api, ctx)


@pytest.mark.case(
    id="UA-2-3-029", chapter="UA-2-3", title="数据源批量_单个与多个",
    preconditions=["目标 datasource alive 且有至少 10 个未注册节点"],
    steps=["选择 1 个节点 batchAdd", "清理", "选择 10 个节点 batchAdd", "验证"],
    expected=["单个和 10 个批量创建均完整"],
)
@pytest.mark.integration
@pytest.mark.destructive
def test_batch_add_single_and_multiple(api, settings, tmp_path_factory, mocker_endpoint):
    ctx = setup_ds_only(api, settings, mocker_endpoint, tmp_path_factory, "UA-2-3-029", nodes=[{"name": "ua23_029_node_", "type": "Double", "count": 11, "writable": True}], namespace_index=2)
    created: list[int] = []
    names: list[str] = []
    try:
        for count, suffix in ((1, "single"), (10, "multiple")):
            nodes = pick_unused_nodes(api, ctx["ds_id"], count=count, namespace_index=2)
            infos = _batch_infos(nodes, ctx["ds_id"], f"{settings.test_prefix}UA-2-3-029-{suffix}")
            result = batch_add_tags(api, infos, conflict_strategy=0)
            assert len(result) == count
            batch_names = [info["tagName"] for info in infos]
            names.extend(batch_names)
            records = _records_for_names(api, batch_names)
            assert len(records) == count
            created.extend(int(r["id"]) for r in records)
            delete_tags_physical(api, created[-count:])
            created = created[:-count]
            names = names[:-count]
    finally:
        if created:
            delete_tags_physical(api, created)
        from tpt_api.datahub import delete_ds_info, change_ds_state
        change_ds_state(api, ctx["ds_id"], False)
        delete_ds_info(api, [ctx["ds_id"]])
        if ctx.get("mocker"):
            from tests.support.mocker_process import stop_mocker
            stop_mocker(ctx["mocker"])


@pytest.mark.case(
    id="UA-2-3-030", chapter="UA-2-3", title="数据源批量_映射隔离",
    preconditions=["两个不同底层节点的数据源 alive"],
    steps=["分别 browse 两个 datasource", "分别 batchAdd", "核对映射和 RT"],
    expected=["dataType、onlyRead、dsId 均按数据源隔离"],
)
@pytest.mark.integration
@pytest.mark.destructive
def test_batch_add_mapping_isolation(api, settings, tmp_path_factory, mocker_endpoint):
    ctx1 = setup_ds_only(api, settings, mocker_endpoint, tmp_path_factory, "UA-2-3-030a", nodes=[{"name": "ua23_030a_", "type": "Double", "count": 1, "writable": True}], namespace_index=2)
    ctx2 = setup_ds_only(api, settings, mocker_endpoint, tmp_path_factory, "UA-2-3-030b", nodes=[{"name": "ua23_030b_", "type": "Int32", "count": 1, "writable": True}], namespace_index=2)
    ids: list[int] = []
    try:
        nodes1 = pick_unused_nodes(api, ctx1["ds_id"], 1, namespace_index=2)
        nodes2 = pick_unused_nodes(api, ctx2["ds_id"], 1, namespace_index=2)
        info1 = _batch_infos(nodes1, ctx1["ds_id"], f"{settings.test_prefix}UA-2-3-030-a")[0]
        info2 = _batch_infos(nodes2, ctx2["ds_id"], f"{settings.test_prefix}UA-2-3-030-b")[0]
        info1["dataType"] = DataTypes["DOUBLE"]
        info2["dataType"] = DataTypes["INT"]
        batch_add_tags(api, [info1], conflict_strategy=0)
        batch_add_tags(api, [info2], conflict_strategy=0)
        rec1 = assert_tag_fields(api, info1["tagName"], dsId=ctx1["ds_id"], dataType=DataTypes["DOUBLE"], onlyRead=True)
        rec2 = assert_tag_fields(api, info2["tagName"], dsId=ctx2["ds_id"], dataType=DataTypes["INT"], onlyRead=True)
        ids = [int(rec1["id"]), int(rec2["id"])]
        _wait_rt(api, info1["tagName"])
        _wait_rt(api, info2["tagName"])
    finally:
        if ids:
            delete_tags_physical(api, ids)
        teardown_ds_tag_mocker(api, ctx2)
        teardown_ds_tag_mocker(api, ctx1)


@pytest.mark.case(
    id="UA-2-3-031", chapter="UA-2-3", title="数据源批量_冲突策略",
    preconditions=["目标 datasource alive"],
    steps=["首次 batchAdd", "strategy=0 重复提交", "strategy=1 修改后提交"],
    expected=["跳过不变，覆盖生效"],
)
@pytest.mark.integration
@pytest.mark.destructive
def test_batch_add_conflict_strategy(api, settings, tmp_path_factory, mocker_endpoint):
    ctx = setup_ds_only(api, settings, mocker_endpoint, tmp_path_factory, "UA-2-3-031", nodes=[{"name": "ua23_031_node_", "type": "Double", "count": 1, "writable": True}], namespace_index=2)
    tag_id = None
    try:
        node = pick_unused_nodes(api, ctx["ds_id"], 1, namespace_index=2)[0]
        info = _batch_infos([node], ctx["ds_id"], f"{settings.test_prefix}UA-2-3-031", unit="kW", desc="original")[0]
        batch_add_tags(api, [info], conflict_strategy=0)
        original = assert_tag_fields(api, info["tagName"], unit="kW", tagDesc="original")
        tag_id = int(original["id"])
        unchanged = dict(info, unit="V", tagDesc="must remain")
        batch_add_tags(api, [unchanged], conflict_strategy=0)
        assert_tag_fields(api, info["tagName"], unit="kW", tagDesc="original")
        batch_add_tags(api, [unchanged], conflict_strategy=1)
        assert_tag_fields(api, info["tagName"], unit="V", tagDesc="must remain", dsId=ctx["ds_id"])
    finally:
        if tag_id:
            delete_tags_physical(api, [tag_id])
        teardown_ds_tag_mocker(api, ctx)


@pytest.mark.case(
    id="UA-2-3-032", chapter="UA-2-3", title="100条完整性",
    preconditions=["目标 datasource alive 且有至少 100 个未注册节点"],
    steps=["browse 100 节点", "一次 batchAdd", "核对数量、唯一性、RT"],
    expected=["恰好 100 条，无重复无遗漏，抽样 RT 可用"],
)
@pytest.mark.integration
@pytest.mark.destructive
def test_batch_add_100_integrity(api, settings, tmp_path_factory, mocker_endpoint):
    ctx = setup_ds_only(api, settings, mocker_endpoint, tmp_path_factory, "UA-2-3-032", nodes=[{"name": "ua23_032_node_", "type": "Double", "count": 100, "change": True, "writable": True}], namespace_index=2)
    ids: list[int] = []
    try:
        nodes = pick_unused_nodes(api, ctx["ds_id"], count=100, namespace_index=2)
        infos = _batch_infos(nodes, ctx["ds_id"], f"{settings.test_prefix}UA-2-3-032")
        result = batch_add_tags(api, infos, conflict_strategy=0)
        assert len(result) == 100
        names = [info["tagName"] for info in infos]
        records = _records_for_names(api, names)
        assert len(records) == 100
        assert len({r["tagName"] for r in records}) == 100
        assert {r["tagBaseName"] for r in records} == {info["tagBaseName"] for info in infos}
        ids = [int(r["id"]) for r in records]
        for name in names[::25]:
            _wait_rt(api, name)
    finally:
        if ids:
            delete_tags_physical(api, ids)
        teardown_ds_tag_mocker(api, ctx)
