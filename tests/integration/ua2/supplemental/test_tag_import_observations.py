"""UA-2-3 supplemental observation tests (no canonical IDs).

Extra coverage beyond the 32 canonical cases.
"""
from __future__ import annotations

import json

import pytest

from tpt_api.datahub import (
    add_tag_group,
    export_tags,
    import_tags_from_file,
    list_tags,
)
from tpt_api.errors import TptAPIError
from tpt_api.types import DataTypes, TagTypes

from tests.support.cleanup import delete_tag_if_exists
from tests.support.polling import wait_until
from tests.support.rt_helpers import get_rt_point
from tests.support.ua2_helpers import (
    setup_ds_and_tag,
    teardown_ds_tag_mocker,
)
from tests.support.ua2_import_helpers import (
    build_import_workbook,
    export_template_bytes,
    import_bytes,
    assert_tag_fields,
    delete_tag_by_name,
    parse_import_result_xlsx,
    tag_type_display,
)


@pytest.mark.integration
@pytest.mark.destructive
@pytest.mark.spec_pending
def test_frequency_boundary(api, settings, tmp_path_factory, mocker_endpoint, record_property):
    """Observe product behavior on frequency edge values."""
    observations = []
    ctx = setup_ds_and_tag(
        api, settings, mocker_endpoint, tmp_path_factory, "UA-2-3-sup-freq",
        tag_base_name="2_ua23_sup_freq_node_1",
        data_type=DataTypes["DOUBLE"],
        tag_type=TagTypes["一次位号"],
        only_read=False,
        nodes=[{"name": "ua23_sup_freq_node_", "type": "Double", "default": 1.0, "writable": True}],
        namespace_index=2,
        cycle=500,
    )
    try:
        tt = tag_type_display(TagTypes["一次位号"])
        template_raw = export_template_bytes(api, ctx["tag_id"])
        freqs = [0, -1, 1, 60, 999999, None]
        rows, tns = [], []
        for i, freq in enumerate(freqs):
            tn = f"{ctx['tag_name']}_freq_{i}"
            tns.append(tn)
            rows.append([tn, "2_ua23_sup_freq_node_1", tt, ctx["ds_name"], "", "DOUBLE",
                         None, None, freq,
                         None, None, None, None, None, None,
                         f"freq {freq}", "Root",
                         "true", "false", None, None, None])
        xlsx = build_import_workbook(template_raw, rows)
        try:
            import_bytes(api, xlsx, conflict_strategy=0, tmp_path_factory=tmp_path_factory)
            for tn in tns:
                page = list_tags(api, page=1, page_size=50, data={"tagName": tn})
                for r in (page.get("records") or []):
                    if r.get("tagName") == tn:
                        observations.append({"tagName": tn, "frequency_saved": r.get("frequency")})
                        record_property("import_freq", observations[-1])
                        delete_tag_if_exists(api, int(r["id"]), tn)
        except TptAPIError as exc:
            record_property("freq_import_error", {"code": exc.code, "msg": exc.msg})
    finally:
        for tn in tns:
            page = list_tags(api, page=1, page_size=50, data={"tagName": tn})
            for r in (page.get("records") or []):
                if r.get("tagName") == tn:
                    delete_tag_if_exists(api, int(r["id"]), tn)
        teardown_ds_tag_mocker(api, ctx)

    pytest.xfail(
        "frequency boundary behavior is product-specific; "
        f"observations: {json.dumps(observations, ensure_ascii=False, default=str)}"
    )


@pytest.mark.integration
@pytest.mark.destructive
@pytest.mark.spec_pending
def test_range_alarm_limits(api, settings, tmp_path_factory, mocker_endpoint, record_property):
    """Observe product behavior on range & alarm limit fields via import."""
    ctx = setup_ds_and_tag(
        api, settings, mocker_endpoint, tmp_path_factory, "UA-2-3-sup-range",
        tag_base_name="2_ua23_sup_range_node_1",
        data_type=DataTypes["DOUBLE"],
        tag_type=TagTypes["一次位号"],
        only_read=False,
        nodes=[{"name": "ua23_sup_range_node_", "type": "Double", "default": 100.0, "writable": True}],
        namespace_index=2,
        cycle=500,
    )
    try:
        tt = tag_type_display(TagTypes["一次位号"])
        template_raw = export_template_bytes(api, ctx["tag_id"])
        tn = f"{ctx['tag_name']}_limits"
        row = [tn, "2_ua23_sup_range_node_1", tt, ctx["ds_name"], "", "DOUBLE",
               None, None, 10,
               180.0, 190.0, 195.0,
               20.0, 10.0, 5.0,
               "limits test", "Root",
               "true", "false", 0.0, 200.0, None]
        xlsx = build_import_workbook(template_raw, [row])
        import_bytes(api, xlsx, conflict_strategy=0, tmp_path_factory=tmp_path_factory)
        rec = assert_tag_fields(api, tn, dsId=ctx["ds_id"])
        record_property("limits", {
            "limitUp": rec.get("limitUp"), "limitUpUp": rec.get("limitUpUp"),
            "limitUpUpUp": rec.get("limitUpUpUp"),
            "limitDown": rec.get("limitDown"), "limitDownDown": rec.get("limitDownDown"),
            "limitDownDownDown": rec.get("limitDownDownDown"),
            "loEu": rec.get("loEu"), "hiEu": rec.get("hiEu"),
        })
    finally:
        delete_tag_by_name(api, tn)
        teardown_ds_tag_mocker(api, ctx)


@pytest.mark.integration
@pytest.mark.destructive
@pytest.mark.spec_pending
def test_base_tag_name_custom(api, settings, tmp_path_factory, mocker_endpoint, record_property):
    """Observe product behavior on custom Base Tag Name via import."""
    ctx = setup_ds_and_tag(
        api, settings, mocker_endpoint, tmp_path_factory, "UA-2-3-sup-base",
        tag_base_name="2_ua23_sup_base_node_1",
        data_type=DataTypes["DOUBLE"],
        tag_type=TagTypes["一次位号"],
        only_read=False,
        nodes=[{"name": "ua23_sup_base_node_", "type": "Double", "default": 1.0, "writable": True}],
        namespace_index=2,
        cycle=500,
    )
    try:
        tt = tag_type_display(TagTypes["一次位号"])
        template_raw = export_template_bytes(api, ctx["tag_id"])
        tn = f"{ctx['tag_name']}_base"
        custom_base = "2_custom_node_sup"
        row = [tn, custom_base, tt, ctx["ds_name"], "", "DOUBLE",
               None, None, 10,
               None, None, None, None, None, None,
               "base name test", "Root",
               "true", "false", None, None, None]
        xlsx = build_import_workbook(template_raw, [row])
        import_bytes(api, xlsx, conflict_strategy=0, tmp_path_factory=tmp_path_factory)
        rec = assert_tag_fields(api, tn, dsId=ctx["ds_id"])
        record_property("base_name", {
            "requested": custom_base,
            "saved": rec.get("tagBaseName"),
        })
    finally:
        delete_tag_by_name(api, tn)
        teardown_ds_tag_mocker(api, ctx)


@pytest.mark.integration
@pytest.mark.destructive
@pytest.mark.spec_pending
def test_push_readonly_combinations(api, settings, tmp_path_factory, mocker_endpoint, record_property):
    """Observe product behavior on push/readonly combinations via import."""
    ctx = setup_ds_and_tag(
        api, settings, mocker_endpoint, tmp_path_factory, "UA-2-3-sup-pr",
        tag_base_name="2_ua23_sup_pr_node_1",
        data_type=DataTypes["DOUBLE"],
        tag_type=TagTypes["一次位号"],
        only_read=False,
        nodes=[{"name": "ua23_sup_pr_node_", "type": "Double", "default": 1.0, "writable": True}],
        namespace_index=2,
        cycle=500,
    )
    try:
        tt = tag_type_display(TagTypes["一次位号"])
        template_raw = export_template_bytes(api, ctx["tag_id"])
        combos = [("true", "false"), ("false", "true"), ("false", "false"), ("true", "true")]
        rows, tns = [], []
        for i, (push, ro) in enumerate(combos):
            tn = f"{ctx['tag_name']}_pr_{i}"
            tns.append(tn)
            rows.append([tn, f"2_ua23_sup_pr_{i}_1", tt, ctx["ds_name"], "", "DOUBLE",
                         None, None, 10,
                         None, None, None, None, None, None,
                         f"push={push} ro={ro}", "Root",
                         push, ro, None, None, None])
        xlsx = build_import_workbook(template_raw, rows)
        import_bytes(api, xlsx, conflict_strategy=0, tmp_path_factory=tmp_path_factory)

        for tn in tns:
            rec = assert_tag_fields(api, tn, dsId=ctx["ds_id"])
            record_property("push_readonly", {
                "tagName": tn,
                "isPush": rec.get("isPush"),
                "onlyRead": rec.get("onlyRead"),
            })
    finally:
        for tn in tns:
            delete_tag_by_name(api, tn)
        teardown_ds_tag_mocker(api, ctx)


@pytest.mark.integration
@pytest.mark.destructive
@pytest.mark.spec_pending
def test_duplicate_headers(api, settings, tmp_path_factory, mocker_endpoint, record_property):
    """Observe product behavior on duplicate column headers."""
    import io
    from openpyxl import Workbook

    ctx = setup_ds_and_tag(
        api, settings, mocker_endpoint, tmp_path_factory, "UA-2-3-sup-duphead",
        tag_base_name="2_ua23_sup_duphead_node_1",
        data_type=DataTypes["DOUBLE"],
        tag_type=TagTypes["一次位号"],
        only_read=False,
        nodes=[{"name": "ua23_sup_duphead_node_", "type": "Double", "default": 1.0, "writable": True}],
        namespace_index=2,
        cycle=500,
    )
    try:
        out_wb = Workbook()
        out_ws = out_wb.active
        dup_header = ["Tag Name", "Base Tag Name", "Tag Name"]
        out_ws.append(dup_header)
        out_ws.append([f"{ctx['tag_name']}_dup", "2_ua23_sup_duphead_node_1", "ignored"])
        buf = io.BytesIO()
        out_wb.save(buf)
        out_wb.close()
        xlsx = buf.getvalue()

        try:
            result = import_bytes(api, xlsx, conflict_strategy=0, tmp_path_factory=tmp_path_factory)
            record_property("duplicate_header_result", "accepted")
        except TptAPIError as exc:
            record_property("duplicate_header_error", {"code": exc.code, "msg": exc.msg})
    finally:
        delete_tag_by_name(api, f"{ctx['tag_name']}_dup")
        teardown_ds_tag_mocker(api, ctx)


@pytest.mark.integration
@pytest.mark.destructive
@pytest.mark.spec_pending
def test_shuffled_columns(api, settings, tmp_path_factory, mocker_endpoint, record_property):
    """Observe product behavior on shuffled column order."""
    import io
    from openpyxl import Workbook as Wb2
    from openpyxl import load_workbook as lw

    ctx = setup_ds_and_tag(
        api, settings, mocker_endpoint, tmp_path_factory, "UA-2-3-sup-shuf",
        tag_base_name="2_ua23_sup_shuf_node_1",
        data_type=DataTypes["DOUBLE"],
        tag_type=TagTypes["一次位号"],
        only_read=False,
        nodes=[{"name": "ua23_sup_shuf_node_", "type": "Double", "default": 1.0, "writable": True}],
        namespace_index=2,
        cycle=500,
    )
    try:
        template_raw = export_template_bytes(api, ctx["tag_id"])
        wb = lw(io.BytesIO(template_raw), read_only=True, data_only=True)
        ws = wb.active
        header = [cell.value for cell in ws[1]]
        wb.close()
        n = len(header)
        shuffled_indices = list(range(n))
        shuffled_indices[0], shuffled_indices[min(1, n-1)] = shuffled_indices[min(1, n-1)], shuffled_indices[0]
        shuffled_header = [header[i] for i in shuffled_indices]

        out_wb = Wb2()
        out_ws = out_wb.active
        out_ws.append(shuffled_header)
        shuffled_row = [None] * n
        for i, si in enumerate(shuffled_indices):
            shuffled_row[si] = f"{ctx['tag_name']}_shuf" if i == 0 else \
                "2_ua23_sup_shuf_node_1" if i == 1 else ("" if i == 3 else None)
        out_ws.append(shuffled_row)
        buf = io.BytesIO()
        out_wb.save(buf)
        out_wb.close()
        xlsx = buf.getvalue()

        try:
            result = import_bytes(api, xlsx, conflict_strategy=0, tmp_path_factory=tmp_path_factory)
            record_property("shuffled_header_result", "accepted")
        except TptAPIError as exc:
            record_property("shuffled_header_error", {"code": exc.code, "msg": exc.msg})
    finally:
        delete_tag_by_name(api, f"{ctx['tag_name']}_shuf")
        teardown_ds_tag_mocker(api, ctx)
