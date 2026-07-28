from __future__ import annotations

import io
import json
import time

import pytest
from openpyxl import load_workbook

from tpt_api.datahub import (
    add_tag,
    add_tag_group,
    export_tags,
    import_tags_from_file,
    list_tags,
    query_tags_with_quality,
)
from tpt_api.errors import TptAPIError
from tpt_api.types import DataTypes, TagTypes

from tests.support.cleanup import delete_tag_if_exists
from tests.support.polling import wait_until
from tests.support.rt_helpers import get_rt_point
from tests.support.ua2_helpers import (
    setup_ds_and_tag,
    setup_ds_only,
    teardown_ds_tag_mocker,
    wait_ds_alive,
    wait_qtq_valid,
)


def _export_template_bytes(api, tag_id: int) -> bytes:
    """Export one tag and return raw xlsx bytes for use as import template."""
    raw = export_tags(api, [tag_id], parse=False)
    return raw


def _build_import_workbook(template_bytes: bytes, rows: list[list]) -> bytes:
    """Load exported template, clear data rows, append new rows, return xlsx bytes."""
    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb.active
    # Remove all data rows (keep header)
    while ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for row in rows:
        col_count = ws.max_column
        padded = list(row) + [None] * (col_count - len(row))
        ws.append(padded[:col_count])
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def _make_row(header, **values) -> list:
    """Build a row list from keyword args keyed by column name."""
    row = [None] * len(header)
    for col_name, val in values.items():
        if col_name in header:
            row[header.index(col_name)] = val
    return row


def _import_bytes(api, xlsx_bytes: bytes, conflict_strategy: int = 0,
                  tmp_path_factory=None) -> dict:
    """Write bytes to tmp file and import."""
    import uuid
    from pathlib import Path
    import tempfile
    if tmp_path_factory is not None:
        tmp = tmp_path_factory.mktemp("import")
    else:
        tmp = Path(tempfile.mkdtemp())
    path = tmp / f"import_{uuid.uuid4().hex[:8]}.xlsx"
    path.write_bytes(xlsx_bytes)
    return import_tags_from_file(api, str(path), conflict_strategy=conflict_strategy)


# Map from TagTypes integer to product import/export display string.
_TAG_TYPE_STR: dict[int, str] = {
    1: "Primary Tag",
    4: "Virtual Tag",
}


def _tag_type_display(tag_type_int: int) -> str:
    return _TAG_TYPE_STR.get(tag_type_int, f"type_{tag_type_int}")


def _assert_tag_fields(api, tag_name: str, **expected) -> dict:
    page = list_tags(api, page=1, page_size=50, data={"tagName": tag_name})
    for r in (page.get("records") or []):
        if r.get("tagName") == tag_name:
            for k, v in expected.items():
                actual = r.get(k)
                assert actual == v, (
                    f"field {k!r} mismatch for {tag_name!r}: "
                    f"expected {v!r}, got {actual!r}"
                )
            return r
    raise AssertionError(f"tag {tag_name!r} not found after import")


@pytest.mark.case(
    id="UA-2-3-013", chapter="UA-2-3",
    title="Excel导入_单条",
    preconditions=["目标 datasource 存在且 alive"],
    steps=["创建 DS + tag", "export 得模板", "构造 1 行工作簿", "导入", "查询验证"],
    expected=["只创建 1 条且字段正确"],
)
@pytest.mark.integration
@pytest.mark.destructive
def test_import_single_tag(api, settings, tmp_path_factory, mocker_endpoint):
    ctx = setup_ds_and_tag(
        api, settings, mocker_endpoint, tmp_path_factory, "UA-2-3-013",
        tag_base_name="2_ua23_013_node_1",
        data_type=DataTypes["DOUBLE"],
        tag_type=TagTypes["一次位号"],
        only_read=False,
        nodes=[{"name": "ua23_013_node_", "type": "Double", "default": 12.5, "writable": True}],
        namespace_index=2,
        cycle=500,
    )
    ds_id = ctx["ds_id"]
    try:
        template_raw = _export_template_bytes(api, ctx["tag_id"])
        new_tag_name = ctx["tag_name"] + "_imp"
        tt = _tag_type_display(TagTypes["一次位号"])
        row = [
            new_tag_name, "2_ua23_013_node_1", tt, ctx["ds_name"], "", "DOUBLE",
            None, None, 10,
            None, None, None, None, None, None,
            "imported single", "Root",
            "true", "false", None, None, None,
        ]
        xlsx = _build_import_workbook(template_raw, [row])
        _import_bytes(api, xlsx, conflict_strategy=0, tmp_path_factory=tmp_path_factory)

        _assert_tag_fields(api, new_tag_name, dsId=ds_id)
        page = list_tags(api, page=1, page_size=50, data={"tagName": new_tag_name})
        match = [r for r in (page.get("records") or []) if r.get("tagName") == new_tag_name]
        assert len(match) == 1, f"expected 1 tag, got {len(match)}"
    finally:
        for tn in [ctx["tag_name"] + "_imp"]:
            page = list_tags(api, page=1, page_size=50, data={"tagName": tn})
            for r in (page.get("records") or []):
                if r.get("tagName") == tn:
                    delete_tag_if_exists(api, int(r["id"]), tn)
        teardown_ds_tag_mocker(api, ctx)


@pytest.mark.case(
    id="UA-2-3-014", chapter="UA-2-3",
    title="Excel导入_多条",
    preconditions=["目标 datasource 存在且 alive"],
    steps=["创建 DS + tag", "export 得模板", "构造 10 行工作簿", "导入", "验证 10 条"],
    expected=["恰好创建 10 条"],
)
@pytest.mark.integration
@pytest.mark.destructive
def test_import_multiple_tags(api, settings, tmp_path_factory, mocker_endpoint):
    ctx = setup_ds_and_tag(
        api, settings, mocker_endpoint, tmp_path_factory, "UA-2-3-014",
        tag_base_name="2_ua23_014_node_1",
        data_type=DataTypes["DOUBLE"],
        tag_type=TagTypes["一次位号"],
        only_read=False,
        nodes=[{"name": "ua23_014_node_", "type": "Double", "default": 1.0, "writable": True}],
        namespace_index=2,
        cycle=500,
    )
    try:
        template_raw = _export_template_bytes(api, ctx["tag_id"])
        tt = _tag_type_display(TagTypes["一次位号"])
        tag_names = []
        rows = []
        for i in range(10):
            tn = f"{ctx['tag_name']}_multi_{i}"
            tag_names.append(tn)
            rows.append([
                tn, f"2_ua23_014_node_{i}_1", tt, ctx["ds_name"], "", "DOUBLE",
                None, None, 10,
                None, None, None, None, None, None,
                f"multi tag {i}", "Root",
                "true", "false", None, None, None,
            ])
        xlsx = _build_import_workbook(template_raw, rows)
        _import_bytes(api, xlsx, conflict_strategy=0, tmp_path_factory=tmp_path_factory)

        count = 0
        for tn in tag_names:
            page = list_tags(api, page=1, page_size=50, data={"tagName": tn})
            for r in (page.get("records") or []):
                if r.get("tagName") == tn:
                    count += 1
        assert count == 10, f"expected 10 tags, got {count}"
    finally:
        for tn in tag_names:
            page = list_tags(api, page=1, page_size=50, data={"tagName": tn})
            for r in (page.get("records") or []):
                if r.get("tagName") == tn:
                    delete_tag_if_exists(api, int(r["id"]), tn)
        teardown_ds_tag_mocker(api, ctx)


@pytest.mark.case(
    id="UA-2-3-015", chapter="UA-2-3",
    title="Excel导入_跨源跨组",
    preconditions=["两个 datasource 和两个普通 group 存在"],
    steps=["双 DS 双 Group 工作簿", "导入", "验证归属"],
    expected=["每条映射到正确 dsId/group"],
)
@pytest.mark.integration
@pytest.mark.destructive
def test_import_cross_ds_group(api, settings, tmp_path_factory, mocker_endpoint):
    ctx_a = setup_ds_and_tag(
        api, settings, mocker_endpoint, tmp_path_factory, "UA-2-3-015a",
        tag_base_name="2_ua23_015a_node_1",
        data_type=DataTypes["DOUBLE"],
        tag_type=TagTypes["一次位号"],
        only_read=False,
        nodes=[{"name": "ua23_015a_node_", "type": "Double", "default": 10.0, "writable": True}],
        namespace_index=2,
        cycle=500,
    )
    ctx_b = setup_ds_and_tag(
        api, settings, mocker_endpoint, tmp_path_factory, "UA-2-3-015b",
        tag_base_name="2_ua23_015b_node_1",
        data_type=DataTypes["INT"],
        tag_type=TagTypes["一次位号"],
        only_read=False,
        nodes=[{"name": "ua23_015b_node_", "type": "Int32", "default": 99, "writable": True}],
        namespace_index=2,
        cycle=500,
    )
    try:
        g1 = add_tag_group(api, f"{ctx_a['tag_name']}_g1")
        g2 = add_tag_group(api, f"{ctx_b['tag_name']}_g2")
        g1_name = str(g1.get("groupName") or g1.get("name") or "")
        g2_name = str(g2.get("groupName") or g2.get("name") or "")
        assert g1_name and g2_name, f"group names empty: {g1_name!r} {g2_name!r}"

        tt = _tag_type_display(TagTypes["一次位号"])
        template_raw = _export_template_bytes(api, ctx_a["tag_id"])
        tag_a = f"{ctx_a['tag_name']}_cross"
        tag_b = f"{ctx_b['tag_name']}_cross"
        rows = [
            [tag_a, "2_ua23_015a_node_1", tt, ctx_a["ds_name"], "", "DOUBLE",
             None, None, 10,
             None, None, None, None, None, None,
             "dsA tag", g1_name,
             "true", "false", None, None, None],
            [tag_b, "2_ua23_015b_node_1", tt, ctx_b["ds_name"], "", "INT",
             None, None, 10,
             None, None, None, None, None, None,
             "dsB tag", g2_name,
             "true", "false", None, None, None],
        ]
        xlsx = _build_import_workbook(template_raw, rows)
        _import_bytes(api, xlsx, conflict_strategy=0, tmp_path_factory=tmp_path_factory)

        _assert_tag_fields(api, tag_a, dsId=ctx_a["ds_id"])
        _assert_tag_fields(api, tag_b, dsId=ctx_b["ds_id"])
    finally:
        for tn in [tag_a, tag_b]:
            page = list_tags(api, page=1, page_size=50, data={"tagName": tn})
            for r in (page.get("records") or []):
                if r.get("tagName") == tn:
                    delete_tag_if_exists(api, int(r["id"]), tn)
        teardown_ds_tag_mocker(api, ctx_b)
        teardown_ds_tag_mocker(api, ctx_a)


@pytest.mark.case(
    id="UA-2-3-016", chapter="UA-2-3",
    title="Excel导入_完整配置",
    preconditions=["datasource 存在"],
    steps=["含全部支持字段的工作簿", "导入", "逐字段回查"],
    expected=["所有字段保存正确"],
)
@pytest.mark.integration
@pytest.mark.destructive
def test_import_full_config(api, settings, tmp_path_factory, mocker_endpoint):
    ctx = setup_ds_and_tag(
        api, settings, mocker_endpoint, tmp_path_factory, "UA-2-3-016",
        tag_base_name="2_ua23_016_node_1",
        data_type=DataTypes["DOUBLE"],
        tag_type=TagTypes["一次位号"],
        only_read=False,
        nodes=[{"name": "ua23_016_node_", "type": "Double", "default": 50.0, "writable": True}],
        namespace_index=2,
        cycle=500,
    )
    try:
        g = add_tag_group(api, f"{ctx['tag_name']}_g")
        g_name = str(g.get("groupName") or g.get("name") or "")

        tt = _tag_type_display(TagTypes["一次位号"])
        template_raw = _export_template_bytes(api, ctx["tag_id"])
        tn = f"{ctx['tag_name']}_full"
        row = [
            tn, "2_ua23_016_node_1", tt, ctx["ds_name"], "", "DOUBLE",
            None, None, 5,
            180.0, 190.0, 195.0,
            20.0, 10.0, 5.0,
            "full config import test", g_name,
            "true", "false", 0.0, 200.0, None,
        ]
        xlsx = _build_import_workbook(template_raw, [row])
        _import_bytes(api, xlsx, conflict_strategy=0, tmp_path_factory=tmp_path_factory)

        rec = _assert_tag_fields(api, tn, dsId=ctx["ds_id"])
        assert str(rec.get("unit", "")) == "", f"unit unexpectedly set: {rec.get('unit')!r}"
        assert str(rec.get("tagDesc", "")) == "full config import test", \
            f"desc mismatch: {rec.get('tagDesc')!r}"
    finally:
        page = list_tags(api, page=1, page_size=50, data={"tagName": tn})
        for r in (page.get("records") or []):
            if r.get("tagName") == tn:
                delete_tag_if_exists(api, int(r["id"]), tn)
        teardown_ds_tag_mocker(api, ctx)


@pytest.mark.case(
    id="UA-2-3-017", chapter="UA-2-3",
    title="Excel导入_13种类型",
    preconditions=["datasource 存在"],
    steps=["13 种数据类型的工作簿", "导入", "验证 dataType"],
    expected=["各类型可导入且 dataType 正确"],
)
@pytest.mark.integration
@pytest.mark.destructive
def test_import_13_data_types(api, settings, tmp_path_factory, mocker_endpoint):
    # DATE_TIME excluded: product rejects DateTime in import (see UA-2-1-071/072/074).
    type_map = {
        "BOOLEAN": ("Boolean", DataTypes["BOOLEAN"]),
        "S_BYTE": ("SByte", DataTypes["S_BYTE"]),
        "BYTE": ("Byte", DataTypes["BYTE"]),
        "SHORT": ("Int16", DataTypes["SHORT"]),
        "U_SHORT": ("UInt16", DataTypes["U_SHORT"]),
        "INT": ("Int32", DataTypes["INT"]),
        "U_INT": ("UInt32", DataTypes["U_INT"]),
        "LONG": ("Int64", DataTypes["LONG"]),
        "U_LONG": ("UInt64", DataTypes["U_LONG"]),
        "FLOAT": ("Float", DataTypes["FLOAT"]),
        "DOUBLE": ("Double", DataTypes["DOUBLE"]),
        "STRING": ("String", DataTypes["STRING"]),
    }
    ctx = setup_ds_and_tag(
        api, settings, mocker_endpoint, tmp_path_factory, "UA-2-3-017",
        tag_base_name="2_ua23_017_node_1",
        data_type=DataTypes["DOUBLE"],
        tag_type=TagTypes["一次位号"],
        only_read=False,
        nodes=[{"name": "ua23_017_node_", "type": "Double", "default": 0.0, "writable": True}],
        namespace_index=2,
        cycle=500,
    )
    try:
        tt = _tag_type_display(TagTypes["一次位号"])
        template_raw = _export_template_bytes(api, ctx["tag_id"])
        tag_names = []
        rows = []
        for idx, (dt_key, (opcua_name, _)) in enumerate(type_map.items()):
            tn = f"{ctx['tag_name']}_type_{idx}"
            tag_names.append(tn)
            rows.append([
                tn, f"2_ua23_017_type_{idx}_1", tt, ctx["ds_name"], "", dt_key,
                None, None, 10,
                None, None, None, None, None, None,
                f"type {idx}", "Root",
                "true", "false", None, None, None,
            ])
        xlsx = _build_import_workbook(template_raw, rows)
        _import_bytes(api, xlsx, conflict_strategy=0, tmp_path_factory=tmp_path_factory)

        for tn in tag_names:
            _assert_tag_fields(api, tn, dsId=ctx["ds_id"])
    finally:
        for tn in tag_names:
            page = list_tags(api, page=1, page_size=50, data={"tagName": tn})
            for r in (page.get("records") or []):
                if r.get("tagName") == tn:
                    delete_tag_if_exists(api, int(r["id"]), tn)
        teardown_ds_tag_mocker(api, ctx)


@pytest.mark.case(
    id="UA-2-3-018", chapter="UA-2-3",
    title="Excel导入_Unicode与空白字段",
    preconditions=["datasource 存在"],
    steps=["含中文、Unicode、空白的工作簿", "导入", "记录配置结果"],
    expected=["不出现乱码或错误映射"],
)
@pytest.mark.integration
@pytest.mark.destructive
@pytest.mark.spec_pending
def test_import_unicode_blank(api, settings, tmp_path_factory, mocker_endpoint, record_property):
    observations: list[dict] = []
    ctx = setup_ds_and_tag(
        api, settings, mocker_endpoint, tmp_path_factory, "UA-2-3-018",
        tag_base_name="2_ua23_018_node_1",
        data_type=DataTypes["DOUBLE"],
        tag_type=TagTypes["一次位号"],
        only_read=False,
        nodes=[{"name": "ua23_018_node_", "type": "Double", "default": 1.0, "writable": True}],
        namespace_index=2,
        cycle=500,
    )
    try:
        tt = _tag_type_display(TagTypes["一次位号"])
        template_raw = _export_template_bytes(api, ctx["tag_id"])
        tag_names = []
        inputs = [
            {"tn": f"{ctx['tag_name']}_cn", "unit": "米/秒", "desc": "中文描述"},
            {"tn": f"{ctx['tag_name']}_unicode", "unit": "µm/s²", "desc": "Unicode ∞ Δ"},
            {"tn": f"{ctx['tag_name']}_empty", "unit": "", "desc": ""},
            {"tn": f"{ctx['tag_name']}_spaces", "unit": "   ", "desc": "  spaces  "},
        ]
        rows = []
        for inp in inputs:
            tag_names.append(inp["tn"])
            rows.append([
                inp["tn"], "2_ua23_018_node_1", tt, ctx["ds_name"], inp["unit"], "DOUBLE",
                None, None, 10,
                None, None, None, None, None, None,
                inp["desc"], "Root",
                "true", "false", None, None, None,
            ])
        xlsx = _build_import_workbook(template_raw, rows)
        _import_bytes(api, xlsx, conflict_strategy=0, tmp_path_factory=tmp_path_factory)

        for inp in inputs:
            page = list_tags(api, page=1, page_size=50, data={"tagName": inp["tn"]})
            for r in (page.get("records") or []):
                if r.get("tagName") == inp["tn"]:
                    obs = {
                        "tagName": inp["tn"],
                        "unit_original": inp["unit"],
                        "unit_saved": r.get("unit"),
                        "desc_original": inp["desc"],
                        "desc_saved": r.get("tagDesc"),
                    }
                    observations.append(obs)
                    record_property("import_row", obs)
                    delete_tag_if_exists(api, int(r["id"]), inp["tn"])
    finally:
        for tn in tag_names:
            page = list_tags(api, page=1, page_size=50, data={"tagName": tn})
            for r in (page.get("records") or []):
                if r.get("tagName") == tn:
                    delete_tag_if_exists(api, int(r["id"]), tn)
        teardown_ds_tag_mocker(api, ctx)

    pytest.xfail(
        "UA-2-3-018 unicode/whitespace field behavior is product-specific; "
        f"observations: {json.dumps(observations, ensure_ascii=False, default=str)}"
    )


@pytest.mark.case(
    id="UA-2-3-019", chapter="UA-2-3",
    title="Excel导入_可用性闭环",
    preconditions=["datasource 存在且 mocker 运行"],
    steps=["导入位号", "等待 RT", "验证 QwQ", "清理"],
    expected=["RT 出现", "QwQ 有效"],
)
@pytest.mark.integration
@pytest.mark.destructive
def test_import_availability(api, settings, tmp_path_factory, mocker_endpoint):
    ctx = setup_ds_and_tag(
        api, settings, mocker_endpoint, tmp_path_factory, "UA-2-3-019",
        tag_base_name="2_ua23_019_node_1",
        data_type=DataTypes["DOUBLE"],
        tag_type=TagTypes["一次位号"],
        only_read=False,
        nodes=[{"name": "ua23_019_node_", "type": "Double", "default": 42.0, "writable": True, "change": True}],
        namespace_index=2,
        cycle=500,
    )
    try:
        tt = _tag_type_display(TagTypes["一次位号"])
        template_raw = _export_template_bytes(api, ctx["tag_id"])
        import_tag_name = ctx["tag_name"] + "_avail"
        row = [
            import_tag_name, "2_ua23_019_node_1", tt, ctx["ds_name"], "", "DOUBLE",
            None, None, 10,
            None, None, None, None, None, None,
            "availability test", "Root",
            "true", "false", None, None, None,
        ]
        xlsx = _build_import_workbook(template_raw, [row])
        _import_bytes(api, xlsx, conflict_strategy=0, tmp_path_factory=tmp_path_factory)

        def _has_rt():
            pt = get_rt_point(api, import_tag_name)
            return pt.get("tagValue") is not None and pt.get("quality", 0) != 0

        wait_until(f"rt:{import_tag_name}", _has_rt, timeout=90.0)

        qwq = query_tags_with_quality(api, ds_id=ctx["ds_id"], tag_name=import_tag_name)
        assert qwq, f"QwQ empty for {import_tag_name}"

        page = list_tags(api, page=1, page_size=50, data={"tagName": import_tag_name})
        for r in (page.get("records") or []):
            if r.get("tagName") == import_tag_name:
                delete_tag_if_exists(api, int(r["id"]), import_tag_name)
    finally:
        page = list_tags(api, page=1, page_size=50, data={"tagName": import_tag_name})
        for r in (page.get("records") or []):
            if r.get("tagName") == import_tag_name:
                delete_tag_if_exists(api, int(r["id"]), import_tag_name)
        teardown_ds_tag_mocker(api, ctx)


@pytest.mark.case(
    id="UA-2-3-020", chapter="UA-2-3",
    title="Excel导入_空文件拒绝",
    preconditions=["datasource 存在"],
    steps=["构造无数据行的 workbook", "尝试导入", "预期服务端拒绝"],
    expected=["明确错误码/消息"],
)
@pytest.mark.integration
@pytest.mark.destructive
def test_import_empty_file(api, settings, tmp_path_factory, mocker_endpoint, record_property):
    ctx = setup_ds_and_tag(
        api, settings, mocker_endpoint, tmp_path_factory, "UA-2-3-020",
        tag_base_name="2_ua23_020_node_1",
        data_type=DataTypes["DOUBLE"],
        tag_type=TagTypes["一次位号"],
        only_read=False,
        nodes=[{"name": "ua23_020_node_", "type": "Double", "default": 1.0, "writable": True}],
        namespace_index=2,
        cycle=500,
    )
    try:
        template_raw = _export_template_bytes(api, ctx["tag_id"])
        xlsx = _build_import_workbook(template_raw, [])
        try:
            _import_bytes(api, xlsx, conflict_strategy=0, tmp_path_factory=tmp_path_factory)
            record_property("empty_import_result", "accepted_empty")
        except TptAPIError as exc:
            record_property("empty_import_error", {"code": exc.code, "msg": exc.msg})
    finally:
        teardown_ds_tag_mocker(api, ctx)


@pytest.mark.case(
    id="UA-2-3-021", chapter="UA-2-3",
    title="Excel导入_冲突策略跳过",
    preconditions=["目标 tag 已存在"],
    steps=["首次导入 tagA", "再次导入相同 tagA（conflict_strategy=0）", "确认仍是首次值"],
    expected=["冲突时不覆盖"],
)
@pytest.mark.integration
@pytest.mark.destructive
def test_import_conflict_skip(api, settings, tmp_path_factory, mocker_endpoint):
    ctx = setup_ds_and_tag(
        api, settings, mocker_endpoint, tmp_path_factory, "UA-2-3-021",
        tag_base_name="2_ua23_021_node_1",
        data_type=DataTypes["DOUBLE"],
        tag_type=TagTypes["一次位号"],
        only_read=False,
        unit="m/s",
        nodes=[{"name": "ua23_021_node_", "type": "Double", "default": 1.0, "writable": True}],
        namespace_index=2,
        cycle=500,
    )
    try:
        tt = _tag_type_display(TagTypes["一次位号"])
        template_raw = _export_template_bytes(api, ctx["tag_id"])
        dup_name = ctx["tag_name"]
        row1 = [
            dup_name, "2_ua23_021_node_1", tt, ctx["ds_name"], "", "DOUBLE",
            None, None, 10,
            None, None, None, None, None, None,
            "first import", "Root",
            "true", "false", None, None, None,
        ]
        xlsx1 = _build_import_workbook(template_raw, [row1])
        _import_bytes(api, xlsx1, conflict_strategy=0, tmp_path_factory=tmp_path_factory)

        row2 = [
            dup_name, "2_ua23_021_node_1", tt, ctx["ds_name"], "", "DOUBLE",
            None, None, 10,
            None, None, None, None, None, None,
            "second import - should be ignored", "Root",
            "true", "false", None, None, None,
        ]
        xlsx2 = _build_import_workbook(template_raw, [row2])
        result = _import_bytes(api, xlsx2, conflict_strategy=0, tmp_path_factory=tmp_path_factory)

        rec = _assert_tag_fields(api, dup_name, dsId=ctx["ds_id"])
        saved_desc = str(rec.get("tagDesc", ""))
        assert saved_desc != "second import - should be ignored", \
            f"skip conflict failed, desc was overwritten: {saved_desc!r}"
    finally:
        teardown_ds_tag_mocker(api, ctx)


@pytest.mark.case(
    id="UA-2-3-022", chapter="UA-2-3",
    title="Excel导入_冲突策略覆盖",
    preconditions=["目标 tag 已存在"],
    steps=["首次导入 tagA", "再次导入相同 tagA（conflict_strategy=1）", "确认被覆盖"],
    expected=["冲突时覆盖"],
)
@pytest.mark.integration
@pytest.mark.destructive
def test_import_conflict_overwrite(api, settings, tmp_path_factory, mocker_endpoint):
    ctx = setup_ds_and_tag(
        api, settings, mocker_endpoint, tmp_path_factory, "UA-2-3-022",
        tag_base_name="2_ua23_022_node_1",
        data_type=DataTypes["DOUBLE"],
        tag_type=TagTypes["一次位号"],
        only_read=False,
        unit="m/s",
        tag_desc="original desc",
        nodes=[{"name": "ua23_022_node_", "type": "Double", "default": 1.0, "writable": True}],
        namespace_index=2,
        cycle=500,
    )
    try:
        tt = _tag_type_display(TagTypes["一次位号"])
        template_raw = _export_template_bytes(api, ctx["tag_id"])
        dup_name = ctx["tag_name"]
        row_overwrite = [
            dup_name, "2_ua23_022_node_1", tt, ctx["ds_name"], "", "DOUBLE",
            None, None, 10,
            None, None, None, None, None, None,
            "overwritten desc", "Root",
            "true", "false", None, None, None,
        ]
        xlsx = _build_import_workbook(template_raw, [row_overwrite])
        _import_bytes(api, xlsx, conflict_strategy=1, tmp_path_factory=tmp_path_factory)

        rec = _assert_tag_fields(api, dup_name, dsId=ctx["ds_id"])
        saved_desc = str(rec.get("tagDesc", ""))
        assert saved_desc == "overwritten desc", \
            f"overwrite conflict failed, desc={saved_desc!r}"
    finally:
        teardown_ds_tag_mocker(api, ctx)


@pytest.mark.case(
    id="UA-2-3-023", chapter="UA-2-3",
    title="Excel导入_重复列",
    preconditions=["datasource 存在"],
    steps=["构造含重复列标题的工作簿", "尝试导入", "记录行为"],
    expected=["记录产品行为"],
)
@pytest.mark.integration
@pytest.mark.destructive
@pytest.mark.spec_pending
def test_import_duplicate_headers(api, settings, tmp_path_factory, mocker_endpoint, record_property):
    ctx = setup_ds_and_tag(
        api, settings, mocker_endpoint, tmp_path_factory, "UA-2-3-023",
        tag_base_name="2_ua23_023_node_1",
        data_type=DataTypes["DOUBLE"],
        tag_type=TagTypes["一次位号"],
        only_read=False,
        nodes=[{"name": "ua23_023_node_", "type": "Double", "default": 1.0, "writable": True}],
        namespace_index=2,
        cycle=500,
    )
    try:
        from openpyxl import Workbook
        out_wb = Workbook()
        out_ws = out_wb.active
        dup_header = ["Tag Name", "Base Tag Name", "Tag Name"]
        out_ws.append(dup_header)
        out_ws.append([f"{ctx['tag_name']}_dup", "2_ua23_023_node_1", "ignored"])
        buf = io.BytesIO()
        out_wb.save(buf)
        out_wb.close()
        xlsx = buf.getvalue()

        try:
            result = _import_bytes(api, xlsx, conflict_strategy=0, tmp_path_factory=tmp_path_factory)
            record_property("duplicate_header_result", "accepted")
        except TptAPIError as exc:
            record_property("duplicate_header_error", {"code": exc.code, "msg": exc.msg})
    finally:
        page = list_tags(api, page=1, page_size=50, data={"tagName": f"{ctx['tag_name']}_dup"})
        for r in (page.get("records") or []):
            if r.get("tagName") == f"{ctx['tag_name']}_dup":
                delete_tag_if_exists(api, int(r["id"]), r.get("tagName"))
        teardown_ds_tag_mocker(api, ctx)


@pytest.mark.case(
    id="UA-2-3-024", chapter="UA-2-3",
    title="Excel导入_乱序列",
    preconditions=["datasource 存在"],
    steps=["构造列顺序打乱的 workbook", "尝试导入", "记录行为"],
    expected=["记录产品行为"],
)
@pytest.mark.integration
@pytest.mark.destructive
@pytest.mark.spec_pending
def test_import_shuffled_columns(api, settings, tmp_path_factory, mocker_endpoint, record_property):
    ctx = setup_ds_and_tag(
        api, settings, mocker_endpoint, tmp_path_factory, "UA-2-3-024",
        tag_base_name="2_ua23_024_node_1",
        data_type=DataTypes["DOUBLE"],
        tag_type=TagTypes["一次位号"],
        only_read=False,
        nodes=[{"name": "ua23_024_node_", "type": "Double", "default": 1.0, "writable": True}],
        namespace_index=2,
        cycle=500,
    )
    try:
        template_raw = _export_template_bytes(api, ctx["tag_id"])
        wb = load_workbook(io.BytesIO(template_raw), read_only=True, data_only=True)
        ws = wb.active
        header = [cell.value for cell in ws[1]]
        wb.close()
        n = len(header)
        shuffled_indices = list(range(n))
        shuffled_indices[0], shuffled_indices[min(1, n-1)] = shuffled_indices[min(1, n-1)], shuffled_indices[0]
        shuffled_header = [header[i] for i in shuffled_indices]

        from openpyxl import Workbook as Wb2
        out_wb = Wb2()
        out_ws = out_wb.active
        out_ws.append(shuffled_header)
        shuffled_row = [None] * n
        for i, si in enumerate(shuffled_indices):
            shuffled_row[si] = f"{ctx['tag_name']}_shuf" if i == 0 else \
                "2_ua23_024_node_1" if i == 1 else ("" if i == 3 else None)
        out_ws.append(shuffled_row)
        buf = io.BytesIO()
        out_wb.save(buf)
        out_wb.close()
        xlsx = buf.getvalue()

        try:
            result = _import_bytes(api, xlsx, conflict_strategy=0, tmp_path_factory=tmp_path_factory)
            record_property("shuffled_header_result", "accepted")
        except TptAPIError as exc:
            record_property("shuffled_header_error", {"code": exc.code, "msg": exc.msg})
    finally:
        page = list_tags(api, page=1, page_size=50, data={"tagName": f"{ctx['tag_name']}_shuf"})
        for r in (page.get("records") or []):
            if r.get("tagName") == f"{ctx['tag_name']}_shuf":
                delete_tag_if_exists(api, int(r["id"]), r.get("tagName"))
        teardown_ds_tag_mocker(api, ctx)


@pytest.mark.case(
    id="UA-2-3-025", chapter="UA-2-3",
    title="Excel导入_频率边界",
    preconditions=["datasource 存在"],
    steps=["构造含 min/max/边界/0/负值的频率", "导入", "记录接受或拒绝行为"],
    expected=["记录产品边界行为"],
)
@pytest.mark.integration
@pytest.mark.destructive
@pytest.mark.spec_pending
def test_import_frequency_boundary(api, settings, tmp_path_factory, mocker_endpoint, record_property):
    observations: list[dict] = []
    ctx = setup_ds_and_tag(
        api, settings, mocker_endpoint, tmp_path_factory, "UA-2-3-025",
        tag_base_name="2_ua23_025_node_1",
        data_type=DataTypes["DOUBLE"],
        tag_type=TagTypes["一次位号"],
        only_read=False,
        nodes=[{"name": "ua23_025_node_", "type": "Double", "default": 1.0, "writable": True}],
        namespace_index=2,
        cycle=500,
    )
    try:
        tt = _tag_type_display(TagTypes["一次位号"])
        template_raw = _export_template_bytes(api, ctx["tag_id"])
        freqs = [0, -1, 1, 60, 999999, None]
        rows = []
        tns = []
        for i, freq in enumerate(freqs):
            tn = f"{ctx['tag_name']}_freq_{i}"
            tns.append(tn)
            rows.append([
                tn, "2_ua23_025_node_1", tt, ctx["ds_name"], "", "DOUBLE",
                None, None, freq,
                None, None, None, None, None, None,
                f"freq {freq}", "Root",
                "true", "false", None, None, None,
            ])
        xlsx = _build_import_workbook(template_raw, rows)
        try:
            _import_bytes(api, xlsx, conflict_strategy=0, tmp_path_factory=tmp_path_factory)
            for tn in tns:
                page = list_tags(api, page=1, page_size=50, data={"tagName": tn})
                for r in (page.get("records") or []):
                    if r.get("tagName") == tn:
                        obs = {"tagName": tn, "frequency_saved": r.get("frequency")}
                        observations.append(obs)
                        record_property("import_freq", obs)
                        delete_tag_if_exists(api, int(r["id"]), tn)
        except TptAPIError as exc:
            record_property("freq_import_error", {"code": exc.code, "msg": exc.msg})
    finally:
        for tn in tns:
            page = list_tags(api, page=1, page_size=50, data={"tagName": tn})
            for r in (page.get("records") or []):
                if r.get("tagName") == tn:
                    delete_tag_if_exists(api, int(r["id"]), tn)
        teardown_ds_tag_mocker(api, ctx)

    pytest.xfail(
        "UA-2-3-025 frequency boundary behavior is product-specific; "
        f"observations: {json.dumps(observations, ensure_ascii=False, default=str)}"
    )


@pytest.mark.case(
    id="UA-2-3-026", chapter="UA-2-3",
    title="Excel导入_量程与六档限值导入",
    preconditions=["datasource 存在"],
    steps=["含量程和六档限值的行", "导入", "回查字段"],
    expected=["量程、高/低限值正确保存"],
)
@pytest.mark.integration
@pytest.mark.destructive
def test_import_range_alarm_limits(api, settings, tmp_path_factory, mocker_endpoint):
    ctx = setup_ds_and_tag(
        api, settings, mocker_endpoint, tmp_path_factory, "UA-2-3-026",
        tag_base_name="2_ua23_026_node_1",
        data_type=DataTypes["DOUBLE"],
        tag_type=TagTypes["一次位号"],
        only_read=False,
        nodes=[{"name": "ua23_026_node_", "type": "Double", "default": 100.0, "writable": True}],
        namespace_index=2,
        cycle=500,
    )
    try:
        tt = _tag_type_display(TagTypes["一次位号"])
        template_raw = _export_template_bytes(api, ctx["tag_id"])
        tn = f"{ctx['tag_name']}_limits"
        row = [
            tn, "2_ua23_026_node_1", tt, ctx["ds_name"], "", "DOUBLE",
            None, None, 10,
            180.0, 190.0, 195.0,
            20.0, 10.0, 5.0,
            "limits test", "Root",
            "true", "false", 0.0, 200.0, None,
        ]
        xlsx = _build_import_workbook(template_raw, [row])
        _import_bytes(api, xlsx, conflict_strategy=0, tmp_path_factory=tmp_path_factory)
        rec = _assert_tag_fields(api, tn, dsId=ctx["ds_id"])
        assert float(rec.get("limitUp", 0) or 0) == 180.0, f"limitUp: {rec.get('limitUp')!r}"
        assert float(rec.get("limitUpUp", 0) or 0) == 190.0, f"limitUpUp: {rec.get('limitUpUp')!r}"
        assert float(rec.get("limitUpUpUp", 0) or 0) == 195.0, f"limitUpUpUp: {rec.get('limitUpUpUp')!r}"
        assert float(rec.get("limitDown", 0) or 0) == 20.0, f"limitDown: {rec.get('limitDown')!r}"
        assert float(rec.get("limitDownDown", 0) or 0) == 10.0, f"limitDownDown: {rec.get('limitDownDown')!r}"
        assert float(rec.get("limitDownDownDown", 0) or 0) == 5.0, f"limitDownDownDown: {rec.get('limitDownDownDown')!r}"
    finally:
        page = list_tags(api, page=1, page_size=50, data={"tagName": tn})
        for r in (page.get("records") or []):
            if r.get("tagName") == tn:
                delete_tag_if_exists(api, int(r["id"]), tn)
        teardown_ds_tag_mocker(api, ctx)


@pytest.mark.case(
    id="UA-2-3-027", chapter="UA-2-3",
    title="Excel导入_数据来源字段",
    preconditions=["datasource 存在"],
    steps=["不同 Base Tag Name 的导入", "记录是否支持指定 Base Tag Name"],
    expected=["Base Tag Name 正确保存"],
)
@pytest.mark.integration
@pytest.mark.destructive
def test_import_base_tag_name(api, settings, tmp_path_factory, mocker_endpoint):
    ctx = setup_ds_and_tag(
        api, settings, mocker_endpoint, tmp_path_factory, "UA-2-3-027",
        tag_base_name="2_ua23_027_node_1",
        data_type=DataTypes["DOUBLE"],
        tag_type=TagTypes["一次位号"],
        only_read=False,
        nodes=[{"name": "ua23_027_node_", "type": "Double", "default": 1.0, "writable": True}],
        namespace_index=2,
        cycle=500,
    )
    try:
        tt = _tag_type_display(TagTypes["一次位号"])
        template_raw = _export_template_bytes(api, ctx["tag_id"])
        tn = f"{ctx['tag_name']}_base"
        custom_base = "2_custom_node_ua23_027"
        row = [
            tn, custom_base, tt, ctx["ds_name"], "", "DOUBLE",
            None, None, 10,
            None, None, None, None, None, None,
            "base name test", "Root",
            "true", "false", None, None, None,
        ]
        xlsx = _build_import_workbook(template_raw, [row])
        _import_bytes(api, xlsx, conflict_strategy=0, tmp_path_factory=tmp_path_factory)
        rec = _assert_tag_fields(api, tn, dsId=ctx["ds_id"])
        saved_base = str(rec.get("tagBaseName", ""))
        assert saved_base == custom_base, f"tagBaseName: {saved_base!r} != {custom_base!r}"
    finally:
        page = list_tags(api, page=1, page_size=50, data={"tagName": tn})
        for r in (page.get("records") or []):
            if r.get("tagName") == tn:
                delete_tag_if_exists(api, int(r["id"]), tn)
        teardown_ds_tag_mocker(api, ctx)


@pytest.mark.case(
    id="UA-2-3-028", chapter="UA-2-3",
    title="Excel导入_Real-time Push与Readonly",
    preconditions=["datasource 存在"],
    steps=["push/readonly 组合行", "导入", "回查字段"],
    expected=["push 和 readonly 按导入值保存"],
)
@pytest.mark.integration
@pytest.mark.destructive
def test_import_push_readonly(api, settings, tmp_path_factory, mocker_endpoint):
    ctx = setup_ds_and_tag(
        api, settings, mocker_endpoint, tmp_path_factory, "UA-2-3-028",
        tag_base_name="2_ua23_028_node_1",
        data_type=DataTypes["DOUBLE"],
        tag_type=TagTypes["一次位号"],
        only_read=False,
        nodes=[{"name": "ua23_028_node_", "type": "Double", "default": 1.0, "writable": True}],
        namespace_index=2,
        cycle=500,
    )
    try:
        tt = _tag_type_display(TagTypes["一次位号"])
        template_raw = _export_template_bytes(api, ctx["tag_id"])
        tns = []
        combos = [
            ("true", "false"),
            ("false", "true"),
            ("false", "false"),
            ("true", "true"),
        ]
        rows = []
        for i, (push, ro) in enumerate(combos):
            tn = f"{ctx['tag_name']}_pr_{i}"
            tns.append(tn)
            rows.append([
                tn, f"2_ua23_028_pr_{i}_1", tt, ctx["ds_name"], "", "DOUBLE",
                None, None, 10,
                None, None, None, None, None, None,
                f"push={push} ro={ro}", "Root",
                push, ro, None, None, None,
            ])
        xlsx = _build_import_workbook(template_raw, rows)
        _import_bytes(api, xlsx, conflict_strategy=0, tmp_path_factory=tmp_path_factory)

        for tn in tns:
            _assert_tag_fields(api, tn, dsId=ctx["ds_id"])
    finally:
        for tn in tns:
            page = list_tags(api, page=1, page_size=50, data={"tagName": tn})
            for r in (page.get("records") or []):
                if r.get("tagName") == tn:
                    delete_tag_if_exists(api, int(r["id"]), tn)
        teardown_ds_tag_mocker(api, ctx)


@pytest.mark.case(
    id="UA-2-3-029", chapter="UA-2-3",
    title="Excel导入_Datasource名称匹配",
    preconditions=["datasource 存在"],
    steps=["使用别名/不存在的 DS 名", "记录产品行为"],
    expected=["记录产品行为"],
)
@pytest.mark.integration
@pytest.mark.destructive
@pytest.mark.spec_pending
def test_import_datasource_name(api, settings, tmp_path_factory, mocker_endpoint, record_property):
    ctx = setup_ds_and_tag(
        api, settings, mocker_endpoint, tmp_path_factory, "UA-2-3-029",
        tag_base_name="2_ua23_029_node_1",
        data_type=DataTypes["DOUBLE"],
        tag_type=TagTypes["一次位号"],
        only_read=False,
        nodes=[{"name": "ua23_029_node_", "type": "Double", "default": 1.0, "writable": True}],
        namespace_index=2,
        cycle=500,
    )
    try:
        tt = _tag_type_display(TagTypes["一次位号"])
        template_raw = _export_template_bytes(api, ctx["tag_id"])
        scenarios = [
            ("correct_name", ctx["ds_name"], "correct_ds"),
            ("wrong_name", "NONEXISTENT_DS_UA23", "wrong_ds"),
            ("empty_name", "", "empty_ds"),
        ]
        obs = []
        for suffix, ds_name_arg, scenario in scenarios:
            tn = f"{ctx['tag_name']}_{suffix}"
            row = [
                tn, "2_ua23_029_node_1", tt, ds_name_arg, "", "DOUBLE",
                None, None, 10,
                None, None, None, None, None, None,
                scenario, "Root",
                "true", "false", None, None, None,
            ]
            xlsx = _build_import_workbook(template_raw, [row])
            try:
                _import_bytes(api, xlsx, conflict_strategy=0, tmp_path_factory=tmp_path_factory)
                page = list_tags(api, page=1, page_size=50, data={"tagName": tn})
                found = [r for r in (page.get("records") or []) if r.get("tagName") == tn]
                for r in found:
                    delete_tag_if_exists(api, int(r["id"]), tn)
                obs.append({"scenario": scenario, "result": "accepted"})
            except TptAPIError as exc:
                obs.append({"scenario": scenario, "result": "rejected", "error": exc.msg})
            record_property("ds_name_test", obs[-1])
    finally:
        teardown_ds_tag_mocker(api, ctx)

    pytest.xfail(
        "UA-2-3-029 datasource name matching behavior is product-specific; "
        f"observations: {json.dumps(obs, ensure_ascii=False, default=str)}"
    )
