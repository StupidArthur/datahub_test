"""Shared helper functions for UA-2-3 import/export tests."""
from __future__ import annotations

import io
import tempfile
import uuid
from pathlib import Path

from openpyxl import load_workbook

from tpt_api.datahub import export_tags, import_tags_from_file, list_tags, delete_tags_physical
from tpt_api.types import TagTypes


# Map from TagTypes integer to product import/export display string.
_TAG_TYPE_STR: dict[int, str] = {
    1: "Primary Tag",
    4: "Virtual Tag",
}


def tag_type_display(tag_type_int: int) -> str:
    return _TAG_TYPE_STR.get(tag_type_int, f"type_{tag_type_int}")


def export_template_bytes(api, tag_id: int) -> bytes:
    """Export one tag and return raw xlsx bytes for use as import template."""
    return export_tags(api, [tag_id], parse=False)


def build_import_workbook(template_bytes: bytes, rows: list) -> bytes:
    """Load exported template, clear data rows, append new rows, return xlsx bytes.

    Rows can be positional lists or dicts keyed by column name.
    """
    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    col_count = len(header)
    while ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for raw_row in rows:
        if isinstance(raw_row, dict):
            row = [None] * col_count
            for k, v in raw_row.items():
                if k in header:
                    row[header.index(k)] = v
        else:
            row = list(raw_row)
        padded = row + [None] * (col_count - len(row))
        ws.append(padded[:col_count])
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def import_bytes(api, xlsx_bytes: bytes, conflict_strategy: int = 0,
                 tmp_path_factory=None) -> dict:
    """Write bytes to tmp file and import."""
    if tmp_path_factory is not None:
        tmp = tmp_path_factory.mktemp("import")
    else:
        tmp = Path(tempfile.mkdtemp())
    path = tmp / f"import_{uuid.uuid4().hex[:8]}.xlsx"
    path.write_bytes(xlsx_bytes)
    return import_tags_from_file(api, str(path), conflict_strategy=conflict_strategy)


def assert_tag_fields(api, tag_name: str, **expected) -> dict:
    """Assert a tag exists with matching fields, return its record."""
    page = list_tags(api, page=1, page_size=50, data={"tagName": tag_name})
    for r in (page.get("records") or []):
        if r.get("tagName") == tag_name:
            for k, v in expected.items():
                actual = r.get(k)
                assert actual == v, (
                    f"field {k!r} mismatch for {tag_name!r}: "
                    f"expected {v!r}, got {actual!r}"
                )
            return r
    raise AssertionError(f"tag {tag_name!r} not found after import")


def delete_tag_by_name(api, tag_name: str):
    """Delete tag by name if it exists."""
    page = list_tags(api, page=1, page_size=50, data={"tagName": tag_name})
    for r in (page.get("records") or []):
        if r.get("tagName") == tag_name:
            delete_tags_physical(api, [int(r["id"])])


def parse_import_result_xlsx(result: dict) -> list[dict]:
    """Parse per-row errors from import result xlsx.

    Returns list of dicts with keys: tag_name, error_msg
    """
    rows = []
    if result.get("response_type") == "xlsx":
        fb = result.get("content", {}).get("file")
        if fb:
            from openpyxl import load_workbook as lw
            wb = lw(io.BytesIO(fb))
            ws = wb.active
            header = [cell.value for cell in ws[1]]
            try:
                err_col = header.index("Error Msg")
                tn_col = header.index("Tag Name")
            except ValueError:
                wb.close()
                return rows
            for r in ws.iter_rows(min_row=2, values_only=True):
                vals = list(r)
                if len(vals) > err_col and vals[err_col]:
                    rows.append({
                        "tag_name": vals[tn_col] if len(vals) > tn_col else "",
                        "error_msg": str(vals[err_col]),
                    })
            wb.close()
    return rows
